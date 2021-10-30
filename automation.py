import openpyxl as xl

from openpyxl.chart import BarChart, Reference, PieChart


def process_workbook(filename, filename2):
    workbook = xl.load_workbook(filename)
    sheet = workbook['Sheet1']
    # cell = sheet['a1']
    # cell = sheet.cell(1, 1)
    # # print(sheet.max_row)
    # # print(cell.value)
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    chart2 = PieChart()
    chart2.add_data(values)
    sheet.add_chart(chart2, 'e20')
    
    workbook.save(filename2)


process_workbook('transactions.xlsx', 'transcationsPie.xlsx')
